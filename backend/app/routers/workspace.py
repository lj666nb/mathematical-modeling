"""
============================================================
工作台路由 — 数据处理 + 可视化 + 论文编辑 (WRK-004/005/006)
============================================================
"""
import csv
import io
import json
import os
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.models import User
from app.routers.auth import get_current_user

router = APIRouter(prefix="/api/workspace", tags=["建模工作台"])

WORKSPACE_DIR = Path("paper_output/workspace")


def _get_user_dir(user_id: int) -> Path:
    d = WORKSPACE_DIR / str(user_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


@router.post("/upload")
async def upload_data(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """上传数据文件（CSV/Excel/JSON）到工作台"""
    ext = Path(file.filename).suffix.lower()
    if ext not in {".csv", ".tsv", ".xlsx", ".xls", ".json"}:
        raise HTTPException(400, "仅支持 CSV/Excel/JSON 格式")

    user_dir = _get_user_dir(current_user.id)
    file_path = user_dir / file.filename
    content = await file.read()
    file_path.write_bytes(content)

    return {
        "filename": file.filename,
        "size": len(content),
        "path": str(file_path),
    }


@router.get("/files")
async def list_files(current_user: User = Depends(get_current_user)):
    """列出工作台已上传的数据文件"""
    user_dir = _get_user_dir(current_user.id)
    if not user_dir.exists():
        return {"files": []}
    files = []
    for p in sorted(user_dir.rglob("*")):
        if p.is_file() and p.suffix.lower() in {".csv", ".tsv", ".xlsx", ".xls", ".json"}:
            files.append({
                "name": p.name,
                "size": p.stat().st_size,
                "ext": p.suffix.lower(),
                "modified": p.stat().st_mtime,
            })
    return {"files": files}


@router.get("/preview/{filename}")
async def preview_data(
    filename: str,
    rows: int = 100,
    current_user: User = Depends(get_current_user),
):
    """预览数据文件内容（前N行）"""
    user_dir = _get_user_dir(current_user.id)
    file_path = user_dir / filename
    if not file_path.exists():
        raise HTTPException(404, "文件不存在")

    ext = file_path.suffix.lower()
    try:
        if ext == ".csv":
            return _preview_csv(file_path, rows)
        elif ext == ".tsv":
            return _preview_csv(file_path, rows, sep="\t")
        elif ext in {".xlsx", ".xls"}:
            return _preview_excel(file_path, rows)
        elif ext == ".json":
            return _preview_json(file_path, rows)
        else:
            raise HTTPException(400, "不支持的文件格式")
    except Exception as e:
        raise HTTPException(500, f"预览失败: {str(e)}")


def _preview_csv(path: Path, rows: int, sep: str = ",") -> dict:
    """预览CSV文件"""
    rows_data = []
    columns = []
    total_rows = 0
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            with path.open("r", encoding=enc) as f:
                reader = csv.reader(f, delimiter=sep)
                columns = next(reader, [])
                columns = [str(c).strip() for c in columns]
                for row in reader:
                    total_rows += 1
                    if len(rows_data) < rows:
                        rows_data.append([str(c).strip() for c in row])
            break
        except Exception:
            continue

    # 推断列类型
    col_types = []
    for ci, col in enumerate(columns):
        vals = [r[ci] for r in rows_data if ci < len(r) and r[ci]]
        nums = sum(1 for v in vals if _is_number(v))
        col_types.append("number" if vals and nums / max(len(vals), 1) >= 0.5 else "string")

    # 基本统计
    stats = {}
    for ci, col in enumerate(columns):
        vals = [float(r[ci].replace(",", "")) for r in rows_data if ci < len(r) and r[ci] and _is_number(r[ci])]
        if vals:
            stats[col] = {
                "min": round(min(vals), 4),
                "max": round(max(vals), 4),
                "mean": round(sum(vals) / len(vals), 4),
                "missing": sum(1 for r in rows_data if ci >= len(r) or not r[ci]),
            }

    return {
        "filename": path.name,
        "columns": columns,
        "col_types": col_types,
        "total_rows": total_rows + len(rows_data),
        "preview_rows": len(rows_data),
        "rows": rows_data,
        "stats": stats,
    }


def _preview_excel(path: Path, rows: int) -> dict:
    """预览Excel文件"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "缺少openpyxl依赖")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        all_rows = []
        columns = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                columns = [str(c) if c is not None else f"Col{j}" for j, c in enumerate(row)]
            else:
                all_rows.append([str(c) if c is not None else "" for c in row])
        sheets[sn] = {
            "columns": columns,
            "total_rows": len(all_rows),
            "rows": all_rows[:rows],
        }
    wb.close()
    return {"filename": path.name, "sheets": sheets}


def _preview_json(path: Path, rows: int) -> dict:
    """预览JSON文件"""
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return {
            "filename": path.name,
            "type": "array",
            "total_rows": len(data),
            "rows": data[:rows],
        }
    return {"filename": path.name, "type": "object", "data": data}


def _is_number(s: str) -> bool:
    try:
        float(s.replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False


@router.post("/clean/{filename}")
async def clean_data(
    filename: str,
    operations: dict = None,
    current_user: User = Depends(get_current_user),
):
    """执行数据清洗操作（去重、填充缺失值、类型转换等）"""
    user_dir = _get_user_dir(current_user.id)
    file_path = user_dir / filename
    if not file_path.exists():
        raise HTTPException(404, "文件不存在")

    if operations is None:
        operations = {}

    try:
        import pandas as pd

        if file_path.suffix.lower() == ".csv":
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)

        op = operations.get("fill_na", "")
        fill_value = operations.get("fill_value", "0")
        if op == "mean":
            for col in df.select_dtypes(include="number").columns:
                df[col] = df[col].fillna(df[col].mean())
        elif op == "median":
            for col in df.select_dtypes(include="number").columns:
                df[col] = df[col].fillna(df[col].median())
        elif op == "mode":
            for col in df.columns:
                df[col] = df[col].fillna(df[col].mode()[0] if not df[col].mode().empty else fill_value)
        elif op == "value":
            df = df.fillna(fill_value)
        elif op == "drop":
            df = df.dropna()

        if operations.get("drop_duplicates"):
            df = df.drop_duplicates()

        out_path = user_dir / f"cleaned_{filename}"
        df.to_csv(out_path, index=False, encoding="utf-8-sig")

        return {
            "original": filename,
            "cleaned": out_path.name,
            "rows_before": len(df),
            "columns": list(df.columns),
            "stats": {
                "missing_after": int(df.isna().sum().sum()),
                "duplicates_removed": operations.get("drop_duplicates", False),
            },
        }
    except ImportError:
        raise HTTPException(500, "缺少pandas依赖")
    except Exception as e:
        raise HTTPException(500, f"清洗失败: {str(e)}")


# ==================== 预置示例数据集 ====================

DEMO_DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "demo"

DEMO_DATASETS = {
    "sample_data": {
        "name": "sample_data.csv",
        "label": "生产区域数据",
        "description": "3个区域 × 3年的生产需求与成本数据，含 demand/cost/capacity/score 字段",
        "rows": 9,
        "columns": ["year", "region", "demand", "cost", "capacity", "score"],
    },
    "q1_oc_curve": {
        "name": "q1_oc_curve.csv",
        "label": "QC抽样曲线",
        "description": "缺陷率与接受概率的关系曲线，301个数据点，适合绘制折线图/散点图",
        "rows": 301,
        "columns": ["defect_rate", "reject_probability", "accept_probability"],
    },
    "q2_all_policy_scores": {
        "name": "q2_all_policy_scores.csv",
        "label": "质检策略评分",
        "description": "6种场景 × 16种策略的质检方案评分，含期望利润/成本/风险等多维度指标",
        "rows": 96,
        "columns": ["scenario", "policy", "inspect_part1", "inspect_part2", "inspect_final",
                     "disassemble_defective", "expected_profit", "expected_cost", "defect_risk",
                     "good_probability", "base_cost", "salvage_value_if_bad"],
    },
}


@router.get("/demo-datasets")
async def list_demo_datasets(current_user: User = Depends(get_current_user)):
    """列出可用的预置示例数据集"""
    datasets = []
    for key, info in DEMO_DATASETS.items():
        file_path = DEMO_DATA_DIR / info["name"]
        available = file_path.exists()
        datasets.append({
            "key": key,
            "label": info["label"],
            "description": info["description"],
            "rows": info["rows"],
            "columns": info["columns"],
            "available": available,
            "size": file_path.stat().st_size if available else 0,
        })
    return {"datasets": datasets}


@router.post("/demo-datasets/{key}/load")
async def load_demo_dataset(
    key: str,
    current_user: User = Depends(get_current_user),
):
    """将预置示例数据集导入到用户工作台"""
    if key not in DEMO_DATASETS:
        raise HTTPException(404, f"未知的示例数据集: {key}")

    info = DEMO_DATASETS[key]
    src = DEMO_DATA_DIR / info["name"]
    if not src.exists():
        raise HTTPException(404, f"示例数据文件不存在: {info['name']}")

    user_dir = _get_user_dir(current_user.id)
    dst = user_dir / info["name"]
    dst.write_bytes(src.read_bytes())

    return {
        "message": f"已导入: {info['label']}",
        "filename": info["name"],
        "label": info["label"],
        "size": dst.stat().st_size,
    }


@router.delete("/files/{filename}")
async def delete_file(
    filename: str,
    current_user: User = Depends(get_current_user),
):
    """删除工作台文件"""
    user_dir = _get_user_dir(current_user.id)
    file_path = user_dir / filename
    if not file_path.exists():
        raise HTTPException(404, "文件不存在")
    file_path.unlink()
    return {"message": f"已删除: {filename}"}
